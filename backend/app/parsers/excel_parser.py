"""
Excel Statement Parser for Bank Exports

Parses Excel exports (.xlsx, .xls) from bank portals with automatic
format detection and data table location.
"""

import io
import re
import logging
from typing import List, Optional, Dict, Any, Tuple
import time

from .base_parser import (
    BaseParser,
    ParsedTransaction,
    ParserResult,
    TransactionType,
    ParserError,
)
from .cleaners import DataCleaner, TransactionCategorizer

logger = logging.getLogger(__name__)


class ExcelStatementParser(BaseParser):
    """
    Parse Excel exports (.xlsx, .xls) from bank portals.
    
    Features:
    - Auto-detect data table location (skip bank logos/headers)
    - Support for multiple sheets
    - Column mapping with format detection
    - Handles merged cells and formatting
    """
    
    PARSER_NAME = "ExcelStatementParser"
    PARSER_VERSION = "1.0.0"
    SUPPORTED_FORMATS = [".xlsx", ".xls", ".xlsm"]
    
    # Column keywords for detection
    COLUMN_KEYWORDS = {
        'date': ['date', 'dt', 'txn date', 'transaction date', 'value date', 'posting date'],
        'description': ['description', 'narration', 'particulars', 'details', 'remarks', 'transaction description'],
        'withdrawal': ['withdrawal', 'debit', 'dr', 'debit amount', 'withdrawal amt', 'dr amount'],
        'deposit': ['deposit', 'credit', 'cr', 'credit amount', 'deposit amt', 'cr amount'],
        'amount': ['amount', 'transaction amount', 'txn amount', 'value'],
        'balance': ['balance', 'closing balance', 'running balance', 'available balance'],
        'reference': ['reference', 'ref no', 'cheque no', 'chq no', 'transaction id', 'utr'],
        'type': ['type', 'transaction type', 'dr/cr', 'cr/dr'],
    }
    
    def __init__(self):
        """Initialize the Excel parser."""
        super().__init__()
        self.cleaner = DataCleaner()
        self.categorizer = TransactionCategorizer()
    
    def can_parse(self, file_content: bytes, filename: str) -> bool:
        """Check if this parser can handle the file."""
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        if ext not in ['xlsx', 'xls', 'xlsm']:
            return False
        
        # Check for Excel magic bytes
        # XLSX (ZIP format)
        if file_content.startswith(b'PK'):
            return True
        # XLS (OLE format)
        if file_content.startswith(b'\xd0\xcf\x11\xe0'):
            return True
        
        return False
    
    def parse(self, file_content: bytes, filename: str) -> ParserResult:
        """
        Parse an Excel statement file.
        
        Args:
            file_content: Raw Excel bytes
            filename: Original filename
            
        Returns:
            ParserResult with parsed transactions
        """
        start_time = time.time()
        result = self._create_result(filename)
        
        try:
            # Determine file type
            is_xlsx = filename.lower().endswith(('.xlsx', '.xlsm'))
            
            if is_xlsx:
                transactions = self._parse_xlsx(file_content, result)
            else:
                transactions = self._parse_xls(file_content, result)
            
            # Add transactions to result
            for txn in transactions:
                if txn:
                    if not txn.category:
                        txn.category = self.categorizer.categorize(txn.description)
                    result.transactions.append(txn)
            
            # Remove duplicates
            unique_txns, duplicates = self.cleaner.remove_duplicates(result.transactions)
            if duplicates:
                result.add_warning(f"Removed {len(duplicates)} duplicate transactions")
                result.transactions = unique_txns
            
            result.parsing_time_ms = (time.time() - start_time) * 1000
            
            self.logger.info(
                f"Parsed {len(result.transactions)} transactions from {filename}"
            )
            
            return result
            
        except Exception as e:
            self.logger.error(f"Excel parsing error: {e}", exc_info=True)
            result.add_error(f"Excel parsing failed: {str(e)}")
            result.parsing_time_ms = (time.time() - start_time) * 1000
            return result
    
    def _parse_xlsx(self, file_content: bytes, result: ParserResult) -> List[ParsedTransaction]:
        """Parse XLSX format using openpyxl."""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.logger.warning("openpyxl not installed, trying pandas")
            return self._parse_with_pandas(file_content, result)
        
        transactions = []
        
        try:
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Skip empty sheets
                if sheet.max_row is None or sheet.max_row < 2:
                    continue
                
                self.logger.info(f"Processing sheet: {sheet_name}")
                
                # Find data start (skip logos, headers)
                data_start_row, data_start_col = self._find_data_start_openpyxl(sheet)
                
                if data_start_row is None:
                    result.add_warning(f"No data found in sheet: {sheet_name}")
                    continue
                
                # Extract header row
                headers = self._extract_headers_openpyxl(sheet, data_start_row, data_start_col)
                
                if not headers:
                    result.add_warning(f"No headers found in sheet: {sheet_name}")
                    continue
                
                # Map columns
                column_map = self._map_columns_from_headers(headers, data_start_col)
                
                if not column_map:
                    result.add_warning(f"Could not map columns in sheet: {sheet_name}")
                    continue
                
                # Parse data rows
                sheet_txns = self._parse_sheet_openpyxl(
                    sheet, data_start_row + 1, column_map, result
                )
                transactions.extend(sheet_txns)
                
                # Try to extract bank info from sheet
                self._extract_bank_info_openpyxl(sheet, result)
            
            wb.close()
            return transactions
            
        except Exception as e:
            self.logger.warning(f"openpyxl parsing failed: {e}, trying pandas")
            return self._parse_with_pandas(file_content, result)
    
    def _parse_xls(self, file_content: bytes, result: ParserResult) -> List[ParsedTransaction]:
        """Parse XLS format using xlrd."""
        try:
            import xlrd
        except ImportError:
            self.logger.warning("xlrd not installed, trying pandas")
            return self._parse_with_pandas(file_content, result)
        
        transactions = []
        
        try:
            wb = xlrd.open_workbook(file_contents=file_content)
            
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                
                if sheet.nrows < 2:
                    continue
                
                self.logger.info(f"Processing sheet: {sheet.name}")
                
                # Find data start
                data_start_row, data_start_col = self._find_data_start_xlrd(sheet)
                
                if data_start_row is None:
                    result.add_warning(f"No data found in sheet: {sheet.name}")
                    continue
                
                # Extract headers
                headers = self._extract_headers_xlrd(sheet, data_start_row, data_start_col)
                
                if not headers:
                    continue
                
                # Map columns
                column_map = self._map_columns_from_headers(headers, data_start_col)
                
                if not column_map:
                    continue
                
                # Parse data rows
                sheet_txns = self._parse_sheet_xlrd(
                    sheet, wb.datemode, data_start_row + 1, column_map, result
                )
                transactions.extend(sheet_txns)
            
            return transactions
            
        except Exception as e:
            self.logger.warning(f"xlrd parsing failed: {e}, trying pandas")
            return self._parse_with_pandas(file_content, result)
    
    def _parse_with_pandas(self, file_content: bytes, result: ParserResult) -> List[ParsedTransaction]:
        """Parse Excel using pandas as fallback."""
        import pandas as pd
        
        transactions = []
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(io.BytesIO(file_content))
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                
                if df.empty:
                    continue
                
                # Find header row
                header_row = self._find_header_row_pandas(df)
                
                if header_row is not None:
                    # Re-read with correct header
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=header_row
                    )
                
                # Clean column names
                df.columns = [str(col).strip().lower() for col in df.columns]
                
                # Map columns
                column_map = self._map_columns_from_df(df)
                
                if not column_map:
                    continue
                
                # Parse rows
                for idx, row in df.iterrows():
                    txn = self._parse_row_pandas(row, column_map)
                    if txn:
                        transactions.append(txn)
            
            return transactions
            
        except Exception as e:
            self.logger.error(f"Pandas Excel parsing failed: {e}")
            result.add_error(f"Excel parsing failed: {str(e)}")
            return []
    
    def _find_data_start_openpyxl(self, sheet) -> Tuple[Optional[int], Optional[int]]:
        """Find where actual data starts in an openpyxl sheet."""
        # Look for header row containing date/description keywords
        for row_idx in range(1, min(sheet.max_row + 1, 30)):
            row_values = []
            for col_idx in range(1, min(sheet.max_column + 1, 20)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_values.append(str(cell_value).lower())
            
            # Check if this row looks like a header
            row_text = ' '.join(row_values)
            if self._is_header_row(row_text):
                # Find first non-empty column
                for col_idx in range(1, sheet.max_column + 1):
                    if sheet.cell(row=row_idx, column=col_idx).value:
                        return row_idx, col_idx
        
        return None, None
    
    def _find_data_start_xlrd(self, sheet) -> Tuple[Optional[int], Optional[int]]:
        """Find where actual data starts in an xlrd sheet."""
        for row_idx in range(min(sheet.nrows, 30)):
            row_values = [str(cell.value).lower() for cell in sheet.row(row_idx) if cell.value]
            row_text = ' '.join(row_values)
            
            if self._is_header_row(row_text):
                for col_idx in range(sheet.ncols):
                    if sheet.cell(row_idx, col_idx).value:
                        return row_idx, col_idx
        
        return None, None
    
    def _find_header_row_pandas(self, df) -> Optional[int]:
        """Find header row in pandas DataFrame."""
        import pandas as pd
        
        for row_idx in range(min(len(df), 30)):
            row_values = [str(v).lower() for v in df.iloc[row_idx].values if pd.notna(v)]
            row_text = ' '.join(row_values)
            
            if self._is_header_row(row_text):
                return row_idx
        
        return None
    
    def _is_header_row(self, row_text: str) -> bool:
        """Check if a row looks like a header row."""
        row_text = row_text.lower()
        
        # Must have date keyword
        has_date = any(kw in row_text for kw in self.COLUMN_KEYWORDS['date'])
        
        # Must have description or amount keyword
        has_desc = any(kw in row_text for kw in self.COLUMN_KEYWORDS['description'])
        has_amount = any(kw in row_text for kw in 
                        self.COLUMN_KEYWORDS['amount'] + 
                        self.COLUMN_KEYWORDS['withdrawal'] + 
                        self.COLUMN_KEYWORDS['deposit'])
        
        return has_date and (has_desc or has_amount)
    
    def _extract_headers_openpyxl(self, sheet, row: int, start_col: int) -> Dict[int, str]:
        """Extract header names from openpyxl sheet."""
        headers = {}
        for col_idx in range(start_col, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            if cell_value:
                headers[col_idx] = str(cell_value).strip().lower()
        return headers
    
    def _extract_headers_xlrd(self, sheet, row: int, start_col: int) -> Dict[int, str]:
        """Extract header names from xlrd sheet."""
        headers = {}
        for col_idx in range(start_col, sheet.ncols):
            cell_value = sheet.cell(row, col_idx).value
            if cell_value:
                headers[col_idx] = str(cell_value).strip().lower()
        return headers
    
    def _map_columns_from_headers(self, headers: Dict[int, str], start_col: int) -> Dict[str, int]:
        """Map column indices to field names."""
        column_map = {}
        
        for col_idx, header in headers.items():
            header_lower = header.lower()
            
            for field, keywords in self.COLUMN_KEYWORDS.items():
                if any(kw in header_lower for kw in keywords):
                    if field not in column_map:  # Don't override first match
                        column_map[field] = col_idx
                    break
        
        # Verify minimum required fields
        if 'date' not in column_map:
            return {}
        
        if 'amount' not in column_map and 'withdrawal' not in column_map and 'deposit' not in column_map:
            return {}
        
        return column_map
    
    def _map_columns_from_df(self, df) -> Dict[str, str]:
        """Map DataFrame columns to field names."""
        column_map = {}
        
        for col in df.columns:
            col_lower = str(col).lower()
            
            for field, keywords in self.COLUMN_KEYWORDS.items():
                if any(kw in col_lower for kw in keywords):
                    if field not in column_map:
                        column_map[field] = col
                    break
        
        if 'date' not in column_map:
            return {}
        
        return column_map
    
    def _parse_sheet_openpyxl(
        self,
        sheet,
        start_row: int,
        column_map: Dict[str, int],
        result: ParserResult
    ) -> List[ParsedTransaction]:
        """Parse transactions from openpyxl sheet."""
        transactions = []
        
        for row_idx in range(start_row, sheet.max_row + 1):
            try:
                row_data = {}
                for field, col_idx in column_map.items():
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data[field] = cell_value
                
                txn = self._create_transaction_from_row(row_data)
                if txn:
                    transactions.append(txn)
            except Exception as e:
                result.add_warning(f"Failed to parse row {row_idx}: {str(e)}")
        
        return transactions
    
    def _parse_sheet_xlrd(
        self,
        sheet,
        datemode: int,
        start_row: int,
        column_map: Dict[str, int],
        result: ParserResult
    ) -> List[ParsedTransaction]:
        """Parse transactions from xlrd sheet."""
        import xlrd
        from datetime import datetime
        
        transactions = []
        
        for row_idx in range(start_row, sheet.nrows):
            try:
                row_data = {}
                for field, col_idx in column_map.items():
                    cell = sheet.cell(row_idx, col_idx)
                    
                    # Handle date cells
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            date_tuple = xlrd.xldate_as_tuple(cell.value, datemode)
                            row_data[field] = datetime(*date_tuple).strftime('%Y-%m-%d')
                        except Exception:
                            row_data[field] = cell.value
                    else:
                        row_data[field] = cell.value
                
                txn = self._create_transaction_from_row(row_data)
                if txn:
                    transactions.append(txn)
            except Exception as e:
                result.add_warning(f"Failed to parse row {row_idx}: {str(e)}")
        
        return transactions
    
    def _parse_row_pandas(
        self,
        row,
        column_map: Dict[str, str]
    ) -> Optional[ParsedTransaction]:
        """Parse a pandas DataFrame row into a transaction."""
        import pandas as pd
        
        row_data = {}
        for field, col_name in column_map.items():
            value = row.get(col_name)
            if pd.notna(value):
                row_data[field] = value
        
        return self._create_transaction_from_row(row_data)
    
    def _create_transaction_from_row(
        self,
        row_data: Dict[str, Any]
    ) -> Optional[ParsedTransaction]:
        """Create a ParsedTransaction from row data dictionary."""
        # Extract date
        date_val = row_data.get('date')
        if not date_val:
            return None
        
        # Handle datetime objects
        if hasattr(date_val, 'strftime'):
            clean_date = date_val.strftime('%Y-%m-%d')
        else:
            clean_date = self.cleaner.clean_date(str(date_val))
        
        if not clean_date:
            return None
        
        # Extract description
        description = row_data.get('description', '')
        if description:
            description = self.cleaner.clean_description(str(description))
        
        # Extract amount and type
        amount = None
        txn_type = TransactionType.UNKNOWN
        
        # Check separate withdrawal/deposit columns
        withdrawal = row_data.get('withdrawal')
        deposit = row_data.get('deposit')
        
        if withdrawal is not None or deposit is not None:
            withdrawal_amt = self.cleaner.clean_amount(str(withdrawal)) if withdrawal else None
            deposit_amt = self.cleaner.clean_amount(str(deposit)) if deposit else None
            
            if withdrawal_amt and withdrawal_amt > 0:
                amount = withdrawal_amt
                txn_type = TransactionType.DEBIT
            elif deposit_amt and deposit_amt > 0:
                amount = deposit_amt
                txn_type = TransactionType.CREDIT
        
        # Check single amount column
        if amount is None and 'amount' in row_data:
            amount = self.cleaner.clean_amount(str(row_data['amount']))
            
            # Check type indicator
            type_val = row_data.get('type', '')
            if type_val:
                type_str = str(type_val).upper().strip()
                if type_str in ['DR', 'D', 'DEBIT', '-']:
                    txn_type = TransactionType.DEBIT
                elif type_str in ['CR', 'C', 'CREDIT', '+']:
                    txn_type = TransactionType.CREDIT
            
            if txn_type == TransactionType.UNKNOWN:
                txn_type = self.cleaner.detect_transaction_type(description, amount=amount)
        
        if amount is None or amount == 0:
            return None
        
        # Extract balance
        balance = None
        if 'balance' in row_data and row_data['balance']:
            balance = self.cleaner.clean_amount(str(row_data['balance']))
        
        # Extract reference
        reference = None
        if 'reference' in row_data and row_data['reference']:
            reference = str(row_data['reference']).strip()
        
        if not reference:
            reference = self.cleaner.extract_reference(description)
        
        return self._create_transaction(
            date=clean_date,
            description=description,
            amount=abs(amount),
            transaction_type=txn_type,
            balance=balance,
            reference=reference,
        )
    
    def _extract_bank_info_openpyxl(self, sheet, result: ParserResult) -> None:
        """Try to extract bank information from sheet."""
        # Check first few rows for bank name
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(5, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    text = str(cell_value).upper()
                    
                    bank_names = {
                        'HDFC BANK': 'HDFC',
                        'STATE BANK OF INDIA': 'SBI',
                        'SBI': 'SBI',
                        'ICICI BANK': 'ICICI',
                        'AXIS BANK': 'AXIS',
                        'KOTAK MAHINDRA': 'KOTAK',
                        'IDFC FIRST': 'IDFC FIRST',
                        'YES BANK': 'YES',
                    }
                    
                    for pattern, bank in bank_names.items():
                        if pattern in text:
                            result.bank_name = bank
                            return


# Backward compatibility alias
ExcelParser = ExcelStatementParser
